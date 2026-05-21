import os
import glob
import json
import re
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Directorios y rutas
WORKSPACE_DIR = os.path.dirname(os.path.abspath(__file__))
BASES_DIR = os.path.join(WORKSPACE_DIR, "bases de datos planes y programas")

# Orden lógico de los cursos para el ordenamiento
ORDEN_CURSOS = [
    "1° Básico", "2° Básico", "3° Básico", "4° Básico", 
    "5° Básico", "6° Básico", "7° Básico", "8° Básico",
    "I Medio", "II Medio", "III Medio", "IV Medio",
    "3 Y 4 Medio", "III y IV Medio"
]

def extraer_numero_oa(oa_str):
    if not isinstance(oa_str, str):
        return 999
    numeros = re.findall(r'\d+', oa_str)
    if numeros:
        return int(numeros[0])
    letra = re.findall(r'OA\s+([a-zA-Z])', oa_str)
    if letra:
        return 1000 + ord(letra[0].lower())
    return 9999

def normalizar_asignatura(nombre: str) -> str:
    n = nombre.strip().title()
    if "Ingles" in n or "Inglés" in n:
        return "Inglés"
    if "Musica" in n or "Música" in n:
        return "Música"
    if "Matematica" in n or "Matemática" in n:
        return "Matemática"
    if "Tecnologia" in n or "Tecnología" in n:
        return "Tecnología"
    if "Orientacion" in n or "Orientación" in n:
        return "Orientación"
    if "Educacion" in n or "Educación" in n:
        if "Fisica" in n or "Física" in n:
            return "Educación Física y Salud"
    if "Artes" in n:
        return "Artes Visuales"
    if "Historia" in n:
        return "Historia, Geografía y Ciencias Sociales"
    if "Ciencias" in n:
        return "Ciencias Naturales"
    if "Lenguaje" in n:
        return "Lenguaje"
    return n

def regenerar_individuales():
    print("Iniciando regeneración de Excels individuales offline...")
    json_files = glob.glob(os.path.join(BASES_DIR, "planes_*.json"))
    json_files = [f for f in json_files if "planes_consolidados_master.json" not in os.path.basename(f)]
    
    if not json_files:
        print("Error: No se encontraron archivos JSON en 'bases de datos planes y programas/'")
        return
        
    print(f"Detectados {len(json_files)} archivos JSON para regenerar.")
    
    for file_path in sorted(json_files):
        filename = os.path.basename(file_path)
        print(f"Procesando {filename}...")
        
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                data = json.load(f)
                
            if not isinstance(data, list):
                print(f"  Advertencia: Estructura no válida en {filename}. Se esperaba una lista.")
                continue
                
            lista_objetivos_flatten = []
            
            for curso_data in data:
                curso = curso_data.get("curso", "Desconocido")
                raw_asignatura = curso_data.get("asignatura", "Desconocida")
                asignatura = normalizar_asignatura(raw_asignatura)
                
                objetivos = curso_data.get("objetivos", [])
                for obj in objetivos:
                    eje = obj.get("eje_curricular", "")
                    num_oa = obj.get("num_oa", "")
                    desc_oa = obj.get("descripcion_oa", "")
                    bloom = obj.get("nivel_bloom", "")
                    indicadores_lista = obj.get("indicadores", [])
                    
                    if isinstance(indicadores_lista, list):
                        indicadores_str = "\n".join([f"• {ind.strip()}" for ind in indicadores_lista if ind.strip()])
                    else:
                        indicadores_str = str(indicadores_lista)
                        
                    lista_objetivos_flatten.append({
                        "Curso": curso,
                        "Asignatura": asignatura,
                        "Eje Curricular": eje,
                        "N° de OA": num_oa,
                        "Descripción del OA": desc_oa,
                        "Nivel Cognitivo (Bloom)": bloom,
                        "Indicadores de Evaluación": indicadores_str
                    })
            
            if not lista_objetivos_flatten:
                print(f"  Advertencia: No hay filas de datos en {filename}. Omitiendo.")
                continue
                
            df = pd.DataFrame(lista_objetivos_flatten)
            
            # Categorización robusta de cursos
            cursos_unicos = df['Curso'].unique()
            categorias_completas = list(ORDEN_CURSOS)
            for c in cursos_unicos:
                if pd.notna(c) and c not in categorias_completas:
                    categorias_completas.append(c)
                    
            df['Curso'] = pd.Categorical(df['Curso'], categories=categorias_completas, ordered=True)
            
            df['_OA_num'] = df['N° de OA'].apply(extraer_numero_oa)
            df = df.sort_values(by=['Curso', 'Asignatura', '_OA_num']).reset_index(drop=True)
            df = df.drop(columns=['_OA_num'])
            
            # Nombre de salida de Excel
            excel_filename = filename.replace(".json", ".xlsx")
            output_excel = os.path.join(BASES_DIR, excel_filename)
            
            writer = pd.ExcelWriter(output_excel, engine='openpyxl')
            df.to_excel(writer, sheet_name='Planes Curriculares', index=False)
            
            workbook = writer.book
            worksheet = writer.sheets['Planes Curriculares']
            
            # Estilos visuales premium
            font_family = "Segoe UI"
            header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
            body_font = Font(name=font_family, size=10, color="333333")
            
            header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            zebra_fill = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
            
            thin_border_side = Side(border_style="thin", color="D9D9D9")
            border_all = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
            
            align_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
            align_center = Alignment(horizontal="center", vertical="top", wrap_text=True)
            
            # Cabecera
            worksheet.row_dimensions[1].height = 28
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = align_center
                cell.border = border_all
                
            # Datos
            for row_num in range(2, len(df) + 2):
                use_zebra = (row_num % 2 == 0)
                worksheet.row_dimensions[row_num].height = None
                for col_num in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.font = body_font
                    cell.border = border_all
                    if use_zebra:
                        cell.fill = zebra_fill
                        
                    col_name = df.columns[col_num - 1]
                    if col_name in ["Curso", "Asignatura", "N° de OA", "Nivel Cognitivo (Bloom)"]:
                        cell.alignment = align_center
                    else:
                        cell.alignment = align_left
            
            # Inmovilizar encabezados y activar filtros
            max_col_letter = get_column_letter(len(df.columns))
            worksheet.auto_filter.ref = f"A1:{max_col_letter}{len(df) + 1}"
            worksheet.freeze_panes = 'A2'
            
            column_widths = {
                "Curso": 14,
                "Asignatura": 24,
                "Eje Curricular": 22,
                "N° de OA": 12,
                "Descripción del OA": 52,
                "Nivel Cognitivo (Bloom)": 22,
                "Indicadores de Evaluación": 64
            }
            for col_idx, col_name in enumerate(df.columns, 1):
                col_letter = get_column_letter(col_idx)
                worksheet.column_dimensions[col_letter].width = column_widths.get(col_name, 20)
                
            writer.close()
            print(f"  -> Guardado y formateado exitosamente en: {output_excel}")
            
        except Exception as e:
            print(f"  Error procesando {filename}: {str(e)}")

if __name__ == "__main__":
    regenerar_individuales()
