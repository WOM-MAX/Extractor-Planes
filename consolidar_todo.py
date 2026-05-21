import os
import glob
import json
import re
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Directorios y rutas
WORKSPACE_DIR = os.path.dirname(os.path.abspath(__file__))
BASES_DIR = os.path.join(WORKSPACE_DIR, "bases de datos planes y programas")

# Rutas de salida para el master consolidado
OUTPUT_EXCEL = os.path.join(BASES_DIR, "planes_consolidados_master.xlsx")
OUTPUT_JSON = os.path.join(BASES_DIR, "planes_consolidados_master.json")

# Orden lógico de los cursos para el ordenamiento
ORDEN_CURSOS = [
    "1° Básico", "2° Básico", "3° Básico", "4° Básico", 
    "5° Básico", "6° Básico", "7° Básico", "8° Básico",
    "I Medio", "II Medio", "III Medio", "IV Medio",
    "3 Y 4 Medio", "III y IV Medio"
]

def normalizar_asignatura(nombre: str) -> str:
    """Normaliza y estandariza los nombres de asignaturas para evitar duplicados."""
    n = nombre.strip().title()
    # Casos especiales
    if "Ingles" in n or "Inglés" in n:
        return "Inglés"
    if "Musica" in n or "Música" in n:
        return "Música"
    if "Matematica" in n or "Matemática" in n:
        return "Matemática"
    if "Tecnologia" in n or "Tecnología" in n:
        return "Tecnología"
    if "Orientacion" in n or "Orientación" in n:
        return "Orientación"
    if "Educacion" in n or "Educación" in n:
        if "Fisica" in n or "Física" in n:
            return "Educación Física y Salud"
    if "Artes" in n:
        return "Artes Visuales"
    if "Ciencias" in n:
        return "Ciencias Naturales"
    if "Historia" in n:
        return "Historia, Geografía y Ciencias Sociales"
    if "Lenguaje" in n:
        return "Lenguaje"
    return n

def extraer_numero_oa(oa_str):
    """Extrae el número de OA para ordenación numérica (ej: 'OA 12' -> 12, 'OA a' -> 999)."""
    if not isinstance(oa_str, str):
        return 999
    numeros = re.findall(r'\d+', oa_str)
    if numeros:
        return int(numeros[0])
    # Si es una letra (ej: 'OA a', 'OA b' de habilidades), ordenamos alfabéticamente al final
    letra = re.findall(r'OA\s+([a-zA-Z])', oa_str)
    if letra:
        return 1000 + ord(letra[0].lower())
    return 9999

def consolidar():
    print("Iniciando consolidación total de asignaturas...")
    
    # Buscar todos los JSON de asignaturas individuales
    # Excluimos planes_consolidados_master.json si ya existe
    json_files = glob.glob(os.path.join(BASES_DIR, "planes_*.json"))
    json_files = [f for f in json_files if "planes_consolidados_master.json" not in os.path.basename(f)]
    
    if not json_files:
        print("Error: No se encontraron archivos JSON individuales en 'bases de datos planes y programas/'")
        return
        
    lista_objetivos_flatten = []
    lista_completa_jerarquica = []
    
    print(f"Detectados {len(json_files)} archivos de asignaturas para procesar.")
    
    for file_path in sorted(json_files):
        filename = os.path.basename(file_path)
        print(f"Procesando archivo: {filename}")
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                data = json.load(f)
                
            # data es una lista de cursos
            if not isinstance(data, list):
                print(f"  Advertencia: Estructura no válida en {filename}. Se esperaba una lista.")
                continue
                
            for curso_data in data:
                curso = curso_data.get("curso", "Desconocido")
                raw_asignatura = curso_data.get("asignatura", "Desconocida")
                asignatura = normalizar_asignatura(raw_asignatura)
                
                # Para el JSON jerárquico guardamos una versión limpia
                clean_curso_data = {
                    "curso": curso,
                    "asignatura": asignatura,
                    "objetivos": curso_data.get("objetivos", [])
                }
                lista_completa_jerarquica.append(clean_curso_data)
                
                objetivos = curso_data.get("objetivos", [])
                for obj in objetivos:
                    eje = obj.get("eje_curricular", "")
                    num_oa = obj.get("num_oa", "")
                    desc_oa = obj.get("descripcion_oa", "")
                    bloom = obj.get("nivel_bloom", "")
                    indicadores_lista = obj.get("indicadores", [])
                    
                    if isinstance(indicadores_lista, list):
                        # Limpiar indicadores vacíos y formatear con viñetas
                        indicadores_str = "\n".join([f"• {ind.strip()}" for ind in indicadores_lista if ind.strip()])
                    else:
                        indicadores_str = str(indicadores_lista)
                        
                    lista_objetivos_flatten.append({
                        "Curso": curso,
                        "Asignatura": asignatura,
                        "Eje Curricular": eje,
                        "N° de OA": num_oa,
                        "Descripción del OA": desc_oa,
                        "Nivel Cognitivo (Bloom)": bloom,
                        "Indicadores de Evaluación": indicadores_str
                    })
        except Exception as e:
            print(f"  Error al procesar {filename}: {str(e)}")

    print(f"Total registros planos extraídos: {len(lista_objetivos_flatten)}")
    
    # 1. Guardar el archivo JSON maestro jerárquico
    try:
        with open(OUTPUT_JSON, "w", encoding="utf-8") as jf:
            json.dump(lista_completa_jerarquica, jf, indent=2, ensure_ascii=False)
        print(f"Archivo JSON consolidado maestro guardado en: {OUTPUT_JSON}")
    except Exception as e:
        print(f"Error al escribir el archivo JSON maestro: {str(e)}")

    if not lista_objetivos_flatten:
        print("Error: No hay datos consolidados para exportar.")
        return

    # 2. Generar el DataFrame y ordenar lógicamente
    df = pd.DataFrame(lista_objetivos_flatten)
    
    # Configurar el tipo categórico ordenado para los Cursos
    df['Curso'] = pd.Categorical(df['Curso'], categories=ORDEN_CURSOS, ordered=True)
    
    # Agregar columna temporal para orden numérico de OA
    df['_OA_num'] = df['N° de OA'].apply(extraer_numero_oa)
    
    # Ordenar por: Asignatura -> Curso -> N° de OA
    df = df.sort_values(by=['Asignatura', 'Curso', '_OA_num']).reset_index(drop=True)
    df = df.drop(columns=['_OA_num'])

    # 3. Crear el libro de Excel con openpyxl para diseño premium interactivo
    try:
        writer = pd.ExcelWriter(OUTPUT_EXCEL, engine='openpyxl')
        
        # --- HOJA 1: PLANES CONSOLIDADOS ---
        df.to_excel(writer, sheet_name='Planes Consolidados', index=False)
        
        # --- HOJA 2: RESUMEN Y MÉTRICAS (Tabla Dinámica / Pivote Estática) ---
        # Contar OAs por Asignatura (filas) y Curso (columnas)
        df_resumen = pd.crosstab(df['Asignatura'], df['Curso'], margins=True, margins_name='Total General')
        
        # Reordenar las columnas del resumen para seguir el orden lógico más el Total General
        columnas_resumen = [c for c in ORDEN_CURSOS if c in df_resumen.columns] + ['Total General']
        df_resumen = df_resumen[columnas_resumen]
        df_resumen.to_excel(writer, sheet_name='Resumen y Métricas')
        
        workbook = writer.book
        
        # --- ESTILIZAR HOJA 1: PLANES CONSOLIDADOS ---
        ws1 = writer.sheets['Planes Consolidados']
        
        # Estilos visuales premium
        font_family = "Segoe UI"
        header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
        body_font = Font(name=font_family, size=10, color="333333")
        total_font = Font(name=font_family, size=10, bold=True, color="1F4E79")
        
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        zebra_fill = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
        
        thin_border_side = Side(border_style="thin", color="D9D9D9")
        border_all = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
        
        align_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
        align_center = Alignment(horizontal="center", vertical="top", wrap_text=True)
        
        # Cabecera de la tabla
        ws1.row_dimensions[1].height = 28
        for col_num in range(1, len(df.columns) + 1):
            cell = ws1.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = border_all
            
        # Contenido de la tabla
        for row_num in range(2, len(df) + 2):
            use_zebra = (row_num % 2 == 0)
            ws1.row_dimensions[row_num].height = None  # Auto-ajuste de alto
            for col_num in range(1, len(df.columns) + 1):
                cell = ws1.cell(row=row_num, column=col_num)
                cell.font = body_font
                cell.border = border_all
                if use_zebra:
                    cell.fill = zebra_fill
                    
                col_name = df.columns[col_num - 1]
                if col_name in ["Curso", "Asignatura", "N° de OA", "Nivel Cognitivo (Bloom)"]:
                    cell.alignment = align_center
                else:
                    cell.alignment = align_left

        # INTERACTIVIDAD CLAVE:
        # A. Activar Filtros Automáticos en el encabezado
        max_col_letter = get_column_letter(len(df.columns))
        ws1.auto_filter.ref = f"A1:{max_col_letter}{len(df) + 1}"
        
        # B. Inmovilizar la primera fila (Encabezado)
        ws1.freeze_panes = 'A2'
        
        # Ajustar anchos específicos y óptimos de columna
        column_widths = {
            "Curso": 14,
            "Asignatura": 24,
            "Eje Curricular": 22,
            "N° de OA": 12,
            "Descripción del OA": 52,
            "Nivel Cognitivo (Bloom)": 22,
            "Indicadores de Evaluación": 64
        }
        for col_idx, col_name in enumerate(df.columns, 1):
            col_letter = get_column_letter(col_idx)
            ws1.column_dimensions[col_letter].width = column_widths.get(col_name, 20)


        # --- ESTILIZAR HOJA 2: RESUMEN Y MÉTRICAS ---
        ws2 = writer.sheets['Resumen y Métricas']
        ws2.views.sheetView[0].showGridLines = True
        
        # Estilos para tabla resumen
        res_header_fill = PatternFill(start_color="2A52BE", end_color="2A52BE", fill_type="solid")
        total_fill = PatternFill(start_color="E6EEF8", end_color="E6EEF8", fill_type="solid")
        
        double_bottom_border = Border(
            left=thin_border_side, right=thin_border_side,
            top=thin_border_side,
            bottom=Side(border_style="double", color="1F4E79")
        )
        
        # Dar formato a la pestaña resumen
        ws2.row_dimensions[1].height = 26
        ws2.column_dimensions['A'].width = 30 # Columna de asignaturas
        
        # Cabecera de la tabla resumen
        for col_num in range(1, len(df_resumen.columns) + 2):
            cell = ws2.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = res_header_fill
            cell.alignment = align_center
            cell.border = border_all
            
        # Celdas del cuerpo del resumen
        for row_num in range(2, len(df_resumen) + 2):
            ws2.row_dimensions[row_num].height = 22
            is_total_row = (row_num == len(df_resumen) + 1)
            
            for col_num in range(1, len(df_resumen.columns) + 2):
                cell = ws2.cell(row=row_num, column=col_num)
                is_total_col = (col_num == len(df_resumen.columns) + 1)
                
                # Valores numéricos alineados al centro
                if col_num > 1:
                    cell.alignment = align_center
                else:
                    cell.alignment = align_left
                
                # Aplicar fuentes y bordes
                if is_total_row or is_total_col:
                    cell.font = total_font
                    cell.fill = total_fill
                    if is_total_row:
                        cell.border = double_bottom_border
                    else:
                        cell.border = border_all
                else:
                    cell.font = body_font
                    cell.border = border_all
                    
                # Ajustar ancho de columnas de datos del resumen
                if col_num > 1:
                    col_letter = get_column_letter(col_num)
                    ws2.column_dimensions[col_letter].width = 14

        writer.close()
        print(f"¡Consolidación exitosa! Excel maestro guardado en: {OUTPUT_EXCEL}")
        print("El archivo incluye autofiltros, paneles fijos e informe de métricas.")
        
    except Exception as e:
        print(f"Error al escribir el archivo Excel maestro: {str(e)}")

if __name__ == "__main__":
    consolidar()
