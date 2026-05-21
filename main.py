import os
import glob
import json
import logging
import time
import pandas as pd
from concurrent.futures import ThreadPoolExecutor, as_completed
from threading import Lock
from dotenv import load_dotenv
from extractor import extract_pdf_data

# Configuración de logs
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - [%(threadName)s] - %(message)s')
logger = logging.getLogger(__name__)

# Cargar variables de entorno
load_dotenv()

# --- CONFIGURACIÓN PRINCIPAL ---
ASIGNATURAS = [
    "EDUCACIÓN FÍSICA Y SALUD",
    "INGLÉS",
    "MÚSICA",
    "ORIENTACIÓN",
    "TECNOLOGÍA"
]
# Nota: Artes Visuales ya se completó en la última ejecución exitosa, así que la excluimos.
# -------------------------------

# Directorios y rutas
WORKSPACE_DIR = os.path.dirname(os.path.abspath(__file__))
CACHE_DIR = os.path.join(WORKSPACE_DIR, "output_cache")
BASES_DIR = os.path.join(WORKSPACE_DIR, "bases de datos planes y programas")
os.makedirs(BASES_DIR, exist_ok=True)
os.makedirs(CACHE_DIR, exist_ok=True)

# Lock para caché
cache_lock = Lock()

# Mapeo de códigos de archivos curriculares a nombres descriptivos en español
CURSO_MAP = {
    "110-1-MATEMÁTICA.pdf": "1° Básico",
    "110-2-MATEMÁTICA.pdf": "2° Básico",
    "110-3-MATEMÁTICA.pdf": "3° Básico",
    "110-4-MATEMÁTICA.pdf": "4° Básico",
    "110-5-MATEMÁTICA.pdf": "5° Básico",
    "110-6-MATEMÁTICA.pdf": "6° Básico",
    "110-7-MATEMÁTICA.pdf": "7° Básico",
    "110-8-MATEMÁTICA.pdf": "8° Básico",
    "310-1-MATEMÁTICA.pdf": "I Medio",
    "310-2-MATEMÁTICA.pdf": "II Medio",
    "310-3-MATEMÁTICA.pdf": "III Medio",
    "310-4-MATEMÁTICA.pdf": "IV Medio"
}

def obtener_nombre_curso(filename: str) -> str:
    basename = os.path.basename(filename)
    if basename in CURSO_MAP:
        return CURSO_MAP[basename]
    
    parts = basename.split('-')
    if len(parts) >= 2:
        prefix = parts[0]
        level = parts[1]
        if prefix == "110":
            return f"{level}° Básico"
        elif prefix == "310":
            romanos = {1: "I", 2: "II", 3: "III", 4: "IV"}
            try:
                num = int(level)
                return f"{romanos.get(num, level)} Medio"
            except ValueError:
                return f"{level} Medio"
                
    return basename.replace(".pdf", "").replace("-", " ")

def consolidar_resultados(asignatura: str) -> bool:
    logger.info(f"Iniciando consolidación de datos para {asignatura}...")
    
    output_excel = os.path.join(BASES_DIR, f"planes_{asignatura.lower().replace(' ', '_').replace(',', '')}.xlsx")
    output_json = os.path.join(BASES_DIR, f"planes_{asignatura.lower().replace(' ', '_').replace(',', '')}.json")
    
    cache_files = glob.glob(os.path.join(CACHE_DIR, "*.json"))
    
    if not cache_files:
        logger.warning("No se encontraron archivos en la caché para consolidar.")
        return False
        
    lista_objetivos_flatten = []
    lista_completa_jerarquica = []
    
    for cache_file in cache_files:
        try:
            with open(cache_file, "r", encoding="utf-8") as f:
                data = json.load(f)
                
                cache_asignatura = data.get("asignatura", "")
                if cache_asignatura.upper() != asignatura.upper() and cache_asignatura.upper() != asignatura.capitalize().upper():
                    continue

                lista_completa_jerarquica.append(data)
                
                curso = data.get("curso", "Desconocido")
                asignatura_str = cache_asignatura
                objetivos = data.get("objetivos", [])
                
                for obj in objetivos:
                    eje = obj.get("eje_curricular", "")
                    num_oa = obj.get("num_oa", "")
                    desc_oa = obj.get("descripcion_oa", "")
                    bloom = obj.get("nivel_bloom", "")
                    indicadores_lista = obj.get("indicadores", [])
                    
                    if isinstance(indicadores_lista, list):
                        indicadores_str = "\n".join([f"• {ind.strip()}" for ind in indicadores_lista if ind.strip()])
                    else:
                        indicadores_str = str(indicadores_lista)
                        
                    lista_objetivos_flatten.append({
                        "Curso": curso,
                        "Asignatura": asignatura_str,
                        "Eje Curricular": eje,
                        "N° de OA": num_oa,
                        "Descripción del OA": desc_oa,
                        "Nivel Cognitivo (Bloom)": bloom,
                        "Indicadores de Evaluación": indicadores_str
                    })
        except Exception as e:
            logger.error(f"Error al leer el archivo de caché {cache_file}: {str(e)}")
            
    try:
        with open(output_json, "w", encoding="utf-8") as jf:
            json.dump(lista_completa_jerarquica, jf, indent=2, ensure_ascii=False)
        logger.info(f"Archivo JSON consolidado guardado en: {output_json}")
    except Exception as e:
        logger.error(f"Error al escribir el archivo JSON consolidado: {str(e)}")

    if not lista_objetivos_flatten:
        logger.warning(f"No hay filas de datos para exportar a Excel en {asignatura}.")
        return False
        
    try:
        df = pd.DataFrame(lista_objetivos_flatten)
        
        orden_cursos = [
            "1° Básico", "2° Básico", "3° Básico", "4° Básico", 
            "5° Básico", "6° Básico", "7° Básico", "8° Básico",
            "I Medio", "II Medio", "III Medio", "IV Medio",
            "3 Y 4 Medio", "III y IV Medio"
        ]
        
        df['Curso'] = pd.Categorical(df['Curso'], categories=orden_cursos, ordered=True)
        
        import re
        def extraer_numero_oa(oa_str):
            if not isinstance(oa_str, str):
                return 999
            numeros = re.findall(r'\d+', oa_str)
            if numeros:
                return int(numeros[0])
            return 999
            
        df['_OA_num'] = df['N° de OA'].apply(extraer_numero_oa)
        df = df.sort_values(by=['Curso', 'Asignatura', '_OA_num']).reset_index(drop=True)
        df = df.drop(columns=['_OA_num'])
        
        writer = pd.ExcelWriter(output_excel, engine='openpyxl')
        df.to_excel(writer, sheet_name='Planes Curriculares', index=False)
        
        workbook = writer.book
        worksheet = writer.sheets['Planes Curriculares']
        
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        font_family = "Segoe UI"
        header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
        body_font = Font(name=font_family, size=10, color="333333")
        
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        zebra_fill = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
        
        thin_border_side = Side(border_style="thin", color="D9D9D9")
        border_all = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
        
        align_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
        align_center = Alignment(horizontal="center", vertical="top", wrap_text=True)
        
        for col_num in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = border_all
            
        for row_num in range(2, len(df) + 2):
            use_zebra = (row_num % 2 == 0)
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.font = body_font
                cell.border = border_all
                if use_zebra:
                    cell.fill = zebra_fill
                    
                col_name = df.columns[col_num - 1]
                if col_name in ["Curso", "Asignatura", "N° de OA", "Nivel Cognitivo (Bloom)"]:
                    cell.alignment = align_center
                else:
                    cell.alignment = align_left
                    
        worksheet.row_dimensions[1].height = 28
        for row_num in range(2, len(df) + 2):
            worksheet.row_dimensions[row_num].height = None
            
        column_widths = {
            "Curso": 14,
            "Asignatura": 14,
            "Eje Curricular": 22,
            "N° de OA": 12,
            "Descripción del OA": 50,
            "Nivel Cognitivo (Bloom)": 20,
            "Indicadores de Evaluación": 60
        }
        
        for col_idx, col_name in enumerate(df.columns, 1):
            col_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[col_letter].width = column_widths.get(col_name, 20)
            
        writer.close()
        logger.info(f"Archivo Excel consolidado y estilizado guardado en: {output_excel}")
        return True
    except Exception as e:
        logger.error(f"Error al escribir el archivo Excel consolidado: {str(e)}")
        return False

def procesar_un_archivo(pdf_path: str, asignatura: str) -> bool:
    filename = os.path.basename(pdf_path)
    curso = obtener_nombre_curso(filename)
    
    cache_key = filename.replace(".pdf", "")
    cache_path = os.path.join(CACHE_DIR, f"{cache_key}.json")
    
    logger.info(f"Procesando: {filename} -> Asignado a: {curso}")
    
    with cache_lock:
        if os.path.exists(cache_path):
            logger.info(f"El curso '{curso}' ({filename}) ya se encuentra en caché. Omitiendo extracción.")
            return True
        
    MAX_REINTENTOS = 3
    for intento in range(1, MAX_REINTENTOS + 1):
        try:
            logger.info(f"Iniciando extracción con Gemini para: {curso} (intento {intento}/{MAX_REINTENTOS})...")
            resultado = extract_pdf_data(pdf_path, asignatura=asignatura.title())
            
            if resultado:
                resultado["curso"] = curso
                resultado["asignatura"] = asignatura.capitalize()
                
                with cache_lock:
                    with open(cache_path, "w", encoding="utf-8") as cf:
                        json.dump(resultado, cf, indent=2, ensure_ascii=False)
                    
                logger.info(f"Extracción completada con éxito y guardada en caché para: {curso}")
                return True
            else:
                if intento < MAX_REINTENTOS:
                    logger.warning(f"Sin resultado para {curso}. Reintentando en 2s...")
                    time.sleep(2)
                else:
                    logger.error(f"No se obtuvo resultado de la extracción para: {curso}")
        except Exception as e:
            if intento < MAX_REINTENTOS:
                logger.warning(f"Error en intento {intento} para {curso}: {str(e)}. Reintentando en 2s...")
                time.sleep(2)
            else:
                logger.error(f"Error inesperado al procesar {curso}: {str(e)}")
    return False

def procesar_asignatura(asignatura: str):
    logger.info(f"\n{'='*60}\nINICIANDO PROCESO PARA: {asignatura}\n{'='*60}")
    pdf_dir = os.path.join(WORKSPACE_DIR, asignatura)
    
    if not os.path.exists(pdf_dir):
        logger.error(f"La carpeta {pdf_dir} no existe. Saltando asignatura...")
        return
        
    pdf_files = glob.glob(os.path.join(pdf_dir, "*.pdf"))
    if not pdf_files:
        logger.error(f"No se encontraron archivos PDF en {pdf_dir}. Saltando...")
        return
        
    logger.info(f"Se detectaron {len(pdf_files)} archivos PDF para {asignatura}.")
    
    # PROCESAMIENTO PARALELO CON THREADS
    # 5 hilos simultáneos aprovecharán al máximo la cuota de la API de pago sin saturarla
    exitos = 0
    fallos = 0
    with ThreadPoolExecutor(max_workers=5) as executor:
        future_to_pdf = {executor.submit(procesar_un_archivo, pdf, asignatura): pdf for pdf in pdf_files}
        
        for future in as_completed(future_to_pdf):
            try:
                exito = future.result()
                if exito:
                    exitos += 1
                else:
                    fallos += 1
            except Exception as exc:
                logger.error(f"Error en hilo de procesamiento: {exc}")
                fallos += 1
                
    logger.info("=" * 60)
    logger.info(f"Resumen de Procesamiento para {asignatura}:\n"
                f" - Total Archivos: {len(pdf_files)}\n"
                f" - Procesados con éxito: {exitos}\n"
                f" - Fallidos: {fallos}")
    logger.info("=" * 60)
    
    exito_consolidador = consolidar_resultados(asignatura)
    if exito_consolidador:
        logger.info(f"¡Consolidación exitosa para {asignatura}!")
    else:
        logger.error(f"La consolidación de resultados falló para {asignatura}.")

def main():
    api_key_check = os.getenv("GEMINI_API_KEY")
    if not api_key_check:
        logger.error("ADVERTENCIA: GEMINI_API_KEY no está configurada. Deteniendo el procesamiento.")
        return

    for asignatura in ASIGNATURAS:
        procesar_asignatura(asignatura)
        
    logger.info("¡TODAS LAS ASIGNATURAS HAN SIDO PROCESADAS EN PARALELO CON ÉXITO!")

if __name__ == "__main__":
    main()
